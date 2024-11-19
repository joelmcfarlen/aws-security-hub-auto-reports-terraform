import json
import sys
import os
import subprocess
import datetime
import re
import boto3

# Install custom packages to /tmp/ and add to path
subprocess.call('pip install openpyxl pymsteams pandas -t /tmp/ --no-cache-dir'.split(), stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
sys.path.insert(1, '/tmp/')

import pandas as pd
import pymsteams
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles
bold_font = Font(bold=True, size=18)
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

aws_account_ids = [
    os.environ.get('aws_account_1', 'n/a'),
    os.environ.get('aws_account_2', 'n/a'),
    os.environ.get('aws_account_3', 'n/a'),
    os.environ.get('aws_account_4', 'n/a'),
    os.environ.get('aws_account_5', 'n/a')
]

today = datetime.date.today()
curr_day = today.strftime('%d')
curr_month = today.strftime('%m')
curr_year = today.strftime('%Y')

marker = None

def clean_generator_id(generator_id):
    if generator_id.startswith("arn:aws:securityhub:::ruleset/"):
        stripped_id = generator_id.replace("arn:aws:securityhub:::ruleset/", "")
        stripped_id = re.sub(r'/rule/\d+(\.\d+)?$', '', stripped_id)
        return stripped_id
    if re.search(r"/v/\d+\.\d+\.\d+/", generator_id):
        match = re.match(r"(.*\/v\/\d+\.\d+\.\d+\/)", generator_id)
        if match:
            return match.group(1)
    if 'security-control' in generator_id:
        return 'security-control'
    return generator_id

# Function to auto-adjust column widths based on cell content
def auto_adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[col_letter].width = adjusted_width

def send_webhook_notification(msg_text, webhook_urls):
    for url in webhook_urls:
        try:
            print(f"Sending message to {url}")
            myTeamsMessage = pymsteams.connectorcard(url)
            myTeamsMessage.title("Security Report Notification")
            myTeamsMessage.color("0000FF")
            myTeamsMessage.text(msg_text)
            myTeamsMessage.send()
        except Exception as e:
            print(f"Failed to send webhook notification to {url}: {e}")

def send_sns_notification(msg_text, sns_topic_arn):
    try:
        print(f"Sending message to SNS topic {sns_topic_arn}")
        sns_client = boto3.client('sns')
        response = sns_client.publish(
            TopicArn=sns_topic_arn,
            Message=msg_text,
            Subject="Security Report Notification"
        )
        print(f"Message sent to SNS topic {sns_topic_arn}. Response: {response}")
    except Exception as e:
        print(f"Failed to send SNS notification: {e}")

def lambda_handler(event, context):
    regions = os.environ['report_regions'].split(',')
    out_bucket = os.environ['out_bucket']
    cust_name = os.environ['cust_name']
    
    # Check for optional environment variables
    webhook_urls = os.environ.get('web_hook_url', '').split(',') if os.environ.get('web_hook_url') else []
    sns_topic_arn = os.environ.get('sns_topic_arn')

    region_identifier = regions[0] if regions else "default-region"
    workbook_name = f"{cust_name}-Security-Hub-Findings-{region_identifier}.xlsx"
    
    headers_row = ['CustomerName', 'Region', 'AccountId', 'Title', 'ProductName', 'Control/Standard', 'Severity', 'ResourceType', 'ResourceId', 'RelatedRequirements']

    try:
        workbook = load_workbook(filename=f"/tmp/{workbook_name}")
        if cust_name in workbook.sheetnames:
            worksheet = workbook[cust_name]
        else:
            workbook.create_sheet(cust_name)
            worksheet = workbook[cust_name]
            worksheet.append(headers_row)
    except FileNotFoundError:
        print(f"File not found. Creating one.")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = cust_name
        worksheet.append(headers_row)

    for region in regions:
        try:
            sh_client = boto3.client('securityhub', region)
            paginator = sh_client.get_paginator('get_findings')
            
            account_filters = [
                {
                    'Value': account_id,
                    'Comparison': 'EQUALS'
                }
                for account_id in aws_account_ids if account_id != 'n/a'
            ]

            page_iterator = paginator.paginate(
                Filters={
                    'AwsAccountId': account_filters,
                    'ComplianceStatus': [
                        {
                            'Value': 'FAILED',
                            'Comparison': 'EQUALS'
                        },
                    ],
                },
                PaginationConfig={
                    'PageSize': 50,
                    'StartingToken': marker
                }
            )

            for page in page_iterator:
                Findings = page['Findings']
                for finding in Findings:
                    try:
                        ProductName = finding['ProductName']
                        AwsAccountId = finding['AwsAccountId']
                        Title = finding['Title']
                        Severity = finding['Severity']['Label']
                        GeneratorId = finding.get('GeneratorId', '')
                        CleanedGeneratorId = clean_generator_id(GeneratorId)
                        RelatedRequirements = ', '.join(finding.get('Compliance', {}).get('RelatedRequirements', []))
                        Resources = finding['Resources']
                        WorkflowStatus = str(finding['Workflow'].get('Status', ''))
                        WorkflowState = finding['WorkflowState']
                        RecordState = finding['RecordState']
                        for resource in Resources:
                            if WorkflowStatus == 'NEW' and RecordState == 'ACTIVE' and WorkflowState == 'NEW':
                                Type = resource['Type']
                                Id = resource['Id']
                                new_row = [cust_name, region, AwsAccountId, Title, ProductName, CleanedGeneratorId, Severity, Type, Id, RelatedRequirements]
                                worksheet.append(new_row)
                    except Exception as e:
                        print(e)

            df = pd.DataFrame(worksheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            df = df.sort_values(by='Severity', ascending=True)
            worksheet.delete_rows(2, worksheet.max_row)
            for index, row in df.iterrows():
                worksheet.append(row.tolist())

            for col_num, cell in enumerate(worksheet[1], 1):
                cell.font = bold_font
                cell.fill = light_blue_fill
                cell.border = thick_border

            worksheet.freeze_panes = worksheet['A2']
            worksheet.auto_filter.ref = worksheet.dimensions

            auto_adjust_column_width(worksheet)

            workbook.save(f"/tmp/{workbook_name}")

            try:
                print('Uploading file to S3')
                s3_client = boto3.client('s3')
                s3_client.upload_file(
                    f"/tmp/{workbook_name}",
                    out_bucket,
                    f"SecurityHubReports/{curr_year}/{curr_month}/{curr_day}/{workbook_name}",
                    ExtraArgs={'ACL': 'bucket-owner-full-control'}
                )
            except Exception as e:
                print(e)

            try:
                msg_text = f"Security Hub Report for {cust_name} is ready in the S3 bucket {out_bucket} in your logging account. You can access it via the console or using the CLI command '$ aws s3 cp s3://{out_bucket}/SecurityHubReports/{curr_year}/{curr_month}/{curr_day}/{workbook_name}' when authenticated to your logging account."

                # Notify via Webhook if configured
                if webhook_urls:
                    send_webhook_notification(msg_text, webhook_urls)
                else:
                    print("No webhook URL configured. Skipping webhook notification.")

                # Notify via SNS if configured
                if sns_topic_arn:
                    send_sns_notification(msg_text, sns_topic_arn)
                else:
                    print("No SNS topic ARN configured. Skipping SNS notification.")
            except Exception as e:
                print(e)

        except Exception as e:
            print(e)
